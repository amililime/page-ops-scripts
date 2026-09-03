"""
Open a Meta/Facebook account from the spreadsheet in a Multilogin browser.

Two modes:

  --manual  (first time or when session expires)
      Opens the Facebook login page inside the Multilogin profile.
      Log in yourself — handles any CAPTCHAs, 2FA, checkpoints.
      Once logged in the script detects it and closes.
      Multilogin persists the session in the profile automatically.

  (no flag)  (every time after)
      Starts the Multilogin profile and opens the profile page directly.
      No login page. If the session has expired it will tell you to
      re-run with --manual.

Usage:
    First time:
        python3 login.py path/to/file.xlsx --account "NOS_PAN_001" --manual

    Every time after:
        python3 login.py path/to/file.xlsx --account "NOS_PAN_001"

Requirements:
    MLX_EMAIL and MLX_PASSWORD environment variables must be set.
    mlx_profiles.json must have an entry for the account.
"""

import argparse
import sys
import time

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

from mlx_context import start_profile_for

SHEET_LAYOUTS = {
    "New VProfiles for BFL": {
        "start_row": 2,
        "name_col":     3,   # C
        "uid_col":      4,   # D
        "password_col": 5,   # E
        "fa_col":       6,   # F
        "email_col":    7,   # G
        "recovery_col": 8,   # H
        "provider_col": 9,   # I
    },
    "BFL Info": {
        "start_row": 3,
        "name_col":     2,   # B
        "uid_col":      5,   # E
        "password_col": 6,   # F
        "fa_col":       7,   # G
        "email_col":    None,
        "recovery_col": 8,   # H
        "provider_col": 9,   # I
    },
}


def normalize(value):
    if value is None:
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def load_credentials(path, account_name, sheet_name):
    try:
        wb = load_workbook(path, data_only=True)
    except FileNotFoundError:
        print(f"Error: file not found at '{path}'.")
        sys.exit(2)

    if sheet_name not in wb.sheetnames:
        print(f"Error: sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        sys.exit(2)

    if sheet_name not in SHEET_LAYOUTS:
        print(f"Error: no layout configured for sheet '{sheet_name}'.")
        sys.exit(2)

    layout = SHEET_LAYOUTS[sheet_name]
    ws = wb[sheet_name]

    for row_idx in range(layout["start_row"], ws.max_row + 1):
        name = normalize(ws.cell(row=row_idx, column=layout["name_col"]).value)
        if name == account_name:
            return {
                "name": name,
                "uid":  normalize(ws.cell(row=row_idx, column=layout["uid_col"]).value),
            }

    print(f"Error: account '{account_name}' not found in sheet '{sheet_name}'.")
    sys.exit(2)


def manual_login(context):
    page = context.pages[0] if context.pages else context.new_page()

    print("Opening Facebook login page...")
    print("→ Log in as usual (email, password, 2FA, any checkpoints).")
    print("→ The script will detect when you're done. Multilogin saves the session automatically.\n")

    page.evaluate("window.location.href = 'https://www.facebook.com/login'")
    page.wait_for_load_state("domcontentloaded")

    timeout = 300
    start = time.time()
    while time.time() - start < timeout:
        if "login" not in page.url and "checkpoint" not in page.url:
            print("Login detected. Session saved in your Multilogin profile.")
            return page
        time.sleep(1)

    print("Timed out waiting for login. Please try again.")
    sys.exit(1)


def restore_session(context):
    page = context.pages[0] if context.pages else context.new_page()

    print("Checking session...")
    page.evaluate("window.location.href = 'https://www.facebook.com/'")
    page.wait_for_load_state("domcontentloaded")
    time.sleep(2)

    if "login" in page.url:
        print("Session has expired.")
        return page, False

    print("Session active.")
    return page, True


def open_account(path, account_name, sheet_name, manual):
    if path:
        creds = load_credentials(path, account_name, sheet_name)
        print(f"Account: {creds['name']}  (UID: {creds['uid']})")
    else:
        creds = {"name": account_name, "uid": None}
        print(f"Account: {account_name}  (no spreadsheet — Multilogin profile only)")

    mlx, started = start_profile_for(account_name)

    with sync_playwright() as p:
        try:
            browser = p.chromium.connect_over_cdp(started.cdp_url)
            context = browser.contexts[0]

            if manual:
                page = manual_login(context)
            else:
                page, valid = restore_session(context)
                if not valid:
                    print("\nRun with --manual to log in first:")
                    print(f'  python3 login.py --account "{account_name}" --manual')
                    sys.exit(1)

            if creds["uid"]:
                profile_url = f"https://www.facebook.com/{creds['uid']}"
                print(f"Opening profile: {profile_url}")
                page.evaluate(f"window.location.href = '{profile_url}'")
                page.wait_for_load_state("domcontentloaded")

            print("\nBrowser is open. Close it when done.")
            try:
                page.wait_for_event("close", timeout=0)
            except Exception:
                pass
        finally:
            mlx.stop_profile(started.profile_id)


def main():
    parser = argparse.ArgumentParser(description="Open a Meta account from the spreadsheet.")
    parser.add_argument("file",      nargs="?", default=None, help="Path to the .xlsx workbook (optional when using mlx_profiles.json)")
    parser.add_argument("--account", required=True, help="Account name (e.g. NOS_PAN_001)")
    parser.add_argument("--sheet",   default="New VProfiles for BFL", help="Sheet name")
    parser.add_argument("--manual",  action="store_true", help="Log in manually (required first time or after session expires)")
    args = parser.parse_args()

    open_account(args.file, args.account, args.sheet, args.manual)


if __name__ == "__main__":
    main()
