"""
Copy a workbook and inject dummy test rows into a named sheet.
Useful for verifying the column layout before tailoring the validation script.

Usage:
    python3 generate_dummy.py path/to/file.xlsx --sheet "BFL Info" --out dummy_output.xlsx
"""

import argparse
import shutil

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# Dummy rows to inject — edit these to match the real column layout
DUMMY_ROWS = [
    {
        "A": "BFL Account 1",
        "B": "JOHN_DOE_001",
        "D": "facebook.com/johnDoe001",
        "E": "12345678901234",
        "F": "PassWord1A",
        "G": "ABCDEFGHIJKLMNOPQRSTUVWXYZ123456",
        "H": "+1234567890",
        "I": "CDC001",
        "J": "2026-12-31",
    },
    {
        "A": "BFL Account 2",
        "B": "JANE_SMITH_002",
        "D": "facebook.com/janeSmith002",
        "E": "98765432109876",
        "F": "SecureAb2C",
        "G": "ZYXWVUTSRQPONMLKJIHGFEDCBA789012",
        "H": "+0987654321",
        "I": "CDC002",
        "J": "2027-06-30",
    },
    {
        "A": "BFL Account 3",
        "B": "MARK_JONES_A003",
        "D": "facebook.com/markJones003",
        "E": "11223344556677",
        "F": "MyPass3Xy1",
        "G": "QWERTYUIOPASDFGHJKLZXCVBNM345678",
        "H": "+1122334455",
        "I": "CDC003",
        "J": "2025-09-15",
    },
]


def inject_dummy(src, sheet_name, out_path, start_row=3):
    shutil.copy2(src, out_path)
    wb = load_workbook(out_path)

    if sheet_name not in wb.sheetnames:
        print(f"Error: sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        return

    ws = wb[sheet_name]

    for i, row_data in enumerate(DUMMY_ROWS):
        row_num = start_row + i
        for col_letter, value in row_data.items():
            col_idx = column_index_from_string(col_letter)
            ws.cell(row=row_num, column=col_idx, value=value)

    wb.save(out_path)
    print(f"Saved {len(DUMMY_ROWS)} dummy rows to: {out_path}")


def main():
    parser = argparse.ArgumentParser(description="Inject dummy rows into a sheet for layout testing.")
    parser.add_argument("file", help="Source .xlsx workbook")
    parser.add_argument("--sheet", default="BFL Info", help="Sheet name to inject into (default: 'BFL Info')")
    parser.add_argument("--out", default="dummy_output.xlsx", help="Output file path (default: dummy_output.xlsx)")
    parser.add_argument("--start-row", type=int, default=3, help="First data row to write (default: 3)")
    args = parser.parse_args()

    inject_dummy(args.file, args.sheet, args.out, args.start_row)


if __name__ == "__main__":
    main()
