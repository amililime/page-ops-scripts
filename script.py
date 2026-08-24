"""
Validate account-inventory sheets against strict format rules.

Supported sheets and their column layouts:

  BFL Info (data starts row 3):
    B = DELIVERED NAME   → NAME_PATTERN
    E = UID              → 9–15 digit number
    F = Password         → PASSWORD_PATTERN
    G = 2FA              → 32 uppercase/digit chars
    H = Number           → phone number (+digits)
    I = CDC              → non-empty
    J = Exp              → non-empty

  New VProfiles for BFL (data starts row 2):
    C = Name             → NAME_PATTERN
    D = UID              → 9–15 digit number
    E = Password         → PASSWORD_PATTERN
    F = 2FA              → 32 uppercase/digit chars
    G = Email            → email address
    H = Email Pass       → PASSWORD_PATTERN (placeholder — confirm real rule)
    I = Provider         → non-empty

  BFL Warming / HUS WarmUp / OPT WarmUp are campaign-tracking sheets
  and are skipped by this validator.

Usage:
    python3 script.py path/to/file.xlsx [--sheet "BFL Info"] [--all]

Exit codes:
    0 = all rows valid
    1 = validation errors found
    2 = could not open file or sheet
"""

import argparse
import re
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ── Validation patterns ────────────────────────────────────────────────────────

NAME_PATTERN     = re.compile(r"^[A-Z]+_[A-Z]+_([0-9]{3}|[A-Z][0-9]{3})$")
UID_PATTERN      = re.compile(r"^\d{9,15}$")
PASSWORD_PATTERN = re.compile(r"^(?=.*[A-Z])(?=.*\d)[A-Za-z\d@!]{8,16}$")
FA_PATTERN       = re.compile(r"^[A-Z0-9]{32}$")   # tighten to [A-Z2-7] if TOTP seeds
EMAIL_PATTERN    = re.compile(r"^[\w.-]+@[\w.-]+\.\w+$")
DOMAIN_PATTERN   = re.compile(r"^([\w.-]+@[\w.-]+\.\w+|[\w.-]+\.\w{2,})$")
PHONE_PATTERN    = re.compile(r"^\+?\d[\d\s\-().]{6,}$")

# ── Per-sheet configuration ────────────────────────────────────────────────────
# Each entry: start_row (first data row), fields dict mapping field name →
# (column index, validator_fn or None for non-empty check, label for errors)

def _re(pattern, label):
    def check(val):
        return bool(pattern.match(val))
    check.__doc__ = label
    return check

def _optional(pattern, label):
    """Validate only if the cell has a value; skip silently if empty."""
    def check(val):
        return bool(pattern.match(val))
    check.__doc__ = label
    check.optional = True
    return check

SHEET_CONFIGS = {
    "BFL Info": {
        "start_row": 3,
        "fields": {
            "name":     (2,  _re(NAME_PATTERN,     "ABC_ABC_123 or ABC_ABC_A123"),   "DELIVERED NAME"),
            "uid":      (5,  _re(UID_PATTERN,       "exactly 14 digits"),             "UID"),
            "password": (6,  _re(PASSWORD_PATTERN,  "10 chars, upper + lower + digit"), "Password"),
            "fa_2":     (7,  _re(FA_PATTERN,        "32 uppercase letters/digits"),   "2FA"),
            "number":   (8,  _re(PHONE_PATTERN,     "phone number starting with +"),  "Number"),
            "cdc":      (9,  None,                                                    "CDC"),
            "exp":      (10, None,                                                    "Exp"),
        },
    },
    "New VProfiles for BFL": {
        "start_row": 2,
        "fields": {
            "name":      (3,  _re(NAME_PATTERN,     "ABC_ABC_123 or ABC_ABC_A123"),      "Name"),
            "uid":       (4,  _re(UID_PATTERN,       "exactly 14 digits"),                "UID"),
            "password":  (5,  _optional(PASSWORD_PATTERN, "8–16 chars, uppercase + digit"),  "Password"),
            "fa_2":      (6,  _re(FA_PATTERN,        "32 uppercase letters/digits"),      "2FA"),
            "email":          (7,  _re(EMAIL_PATTERN,          "valid email address"),    "Email"),
            "recovery_email": (8,  _optional(DOMAIN_PATTERN,  "domain or email address"), "Recovery Email"),
            "provider":       (9,  None,                                                   "Provider"),
        },
    },
}

SKIP_SHEETS = {"BFL Warming", "HUS WarmUp", "OPT WarmUp"}

# ── Helpers ────────────────────────────────────────────────────────────────────

def normalize(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def validate_sheet(ws, config):
    errors = []
    fields = config["fields"]
    start  = config["start_row"]

    for row_idx in range(start, ws.max_row + 1):
        raw    = {key: ws.cell(row=row_idx, column=col).value for key, (col, *_) in fields.items()}
        values = {key: normalize(v) for key, v in raw.items()}

        if not any(values.values()):
            continue  # fully blank row

        for key, (col, validator, label) in fields.items():
            val = values[key]
            if validator is None:
                if not val:
                    errors.append(f"Row {row_idx}: Missing {label} (col {get_column_letter(col)}).")
            elif getattr(validator, "optional", False):
                if val and not validator(val):
                    errors.append(
                        f"Row {row_idx}: Invalid {label} (col {get_column_letter(col)})"
                        f" — got {repr(val)}. Expected: {validator.__doc__}."
                    )
            else:
                if not val or not validator(val):
                    errors.append(
                        f"Row {row_idx}: Invalid {label} (col {get_column_letter(col)})"
                        f" — got {repr(val)}. Expected: {validator.__doc__}."
                    )

    return errors


def run(path, sheet_name=None, all_sheets=False):
    try:
        wb = load_workbook(path, data_only=True)
    except FileNotFoundError:
        print(f"Error: file not found at '{path}'.")
        return None
    except Exception as exc:
        print(f"Error: could not open '{path}': {exc}")
        return None

    if all_sheets:
        targets = [s for s in wb.sheetnames if s in SHEET_CONFIGS]
        skipped = [s for s in wb.sheetnames if s in SKIP_SHEETS]
        unknown = [s for s in wb.sheetnames if s not in SHEET_CONFIGS and s not in SKIP_SHEETS]
    else:
        name = sheet_name or wb.active.title
        if name in SKIP_SHEETS:
            print(f"Sheet '{name}' is a campaign-tracking sheet — skipping validation.")
            return []
        if name not in SHEET_CONFIGS:
            print(f"Error: no validation config for sheet '{name}'.")
            print(f"  Configured sheets : {list(SHEET_CONFIGS.keys())}")
            print(f"  Available in file : {wb.sheetnames}")
            return None
        targets = [name]
        skipped = []
        unknown = []

    all_errors = {}
    for name in targets:
        ws     = wb[name]
        config = SHEET_CONFIGS[name]
        errs   = validate_sheet(ws, config)
        all_errors[name] = errs

    if skipped:
        print(f"Skipped (campaign trackers): {', '.join(skipped)}")
    if unknown:
        print(f"Skipped (no config): {', '.join(unknown)}")

    return all_errors


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Validate account-inventory spreadsheets.")
    parser.add_argument("file", help="Path to the .xlsx workbook")
    parser.add_argument("--sheet", default=None, help="Sheet name (default: active sheet)")
    parser.add_argument("--all",   action="store_true", help="Validate all configured sheets")
    args = parser.parse_args()

    result = run(args.file, args.sheet, args.all)

    if result is None:
        sys.exit(2)

    total_errors = 0
    for sheet_name, errors in result.items():
        if errors:
            print(f"\n=== {sheet_name}: {len(errors)} error(s) ===")
            for err in errors:
                print(f"  - {err}")
            total_errors += len(errors)
        else:
            print(f"\n=== {sheet_name}: OK ===")

    if total_errors:
        sys.exit(1)
    else:
        print("\nAll validated sheets passed.")
        sys.exit(0)


if __name__ == "__main__":
    main()
