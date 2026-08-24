"""
Inspect the structure of every sheet in a workbook.
Prints the first N rows with column letters and values.

Usage:
    python3 inspect_sheets.py path/to/file.xlsx [--rows 5]
"""

import argparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def inspect(path, num_rows=5):
    wb = load_workbook(path, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'=' * 60}")
        print(f"SHEET: {sheet_name}  ({ws.max_row} rows x {ws.max_column} cols)")
        print("=" * 60)

        for row_idx in range(1, min(num_rows + 1, ws.max_row + 1)):
            print(f"  Row {row_idx}:")
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                letter = get_column_letter(col_idx)
                if cell.value is not None:
                    print(f"    Col {letter}: {repr(cell.value)}")
            print()


def main():
    parser = argparse.ArgumentParser(description="Inspect sheet structure of an xlsx workbook.")
    parser.add_argument("file", help="Path to the .xlsx workbook")
    parser.add_argument("--rows", type=int, default=5, help="Number of rows to preview per sheet (default: 5)")
    args = parser.parse_args()

    inspect(args.file, args.rows)


if __name__ == "__main__":
    main()
